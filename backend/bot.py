"""
VMWedding — бэкенд для сбора заявок RSVP + Telegram-бот.

Что делает:
  • Flask-эндпоинт POST /api/rsvp — принимает заявку с формы сайта,
    сохраняет в SQLite и шлёт уведомление в Telegram администраторам.
  • Telegram-бот (long polling в отдельном потоке):
      /start, /menu — меню с кнопками
      «📥 Скачать список гостей (Excel)» — выгрузка xlsx со всеми заявками
      «📊 Статистика» — сводка по ответам

Запуск:  python bot.py
Конфиг:  переменные окружения (см. .env.example)
"""

import os
import io
import json
import time
import sqlite3
import threading
from datetime import datetime, timezone, timedelta

import requests
from flask import Flask, request, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────── Конфиг ───────────────────────────
BOT_TOKEN      = os.environ.get("BOT_TOKEN", "").strip()
ADMIN_CHAT_IDS = [c.strip() for c in os.environ.get("ADMIN_CHAT_IDS", "").split(",") if c.strip()]
DB_PATH        = os.environ.get("DB_PATH", "guests.db")
PORT           = int(os.environ.get("PORT", "8080"))
ALLOW_ORIGIN   = os.environ.get("ALLOW_ORIGIN", "*")
API            = f"https://api.telegram.org/bot{BOT_TOKEN}"

# Часовой пояс Якутска (UTC+9) для меток времени
YAKUTSK_TZ = timezone(timedelta(hours=9))

ATTENDANCE_RU = {
    "yes":   "✅ Да, смогу",
    "maybe": "🤔 Пока не уверен(а)",
    "no":    "❌ К сожалению, не смогу",
}

# ─────────────────────────── База данных ───────────────────────────
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS guests (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                attendance TEXT NOT NULL,
                comment    TEXT,
                created_at TEXT NOT NULL
            )
        """)

def save_guest(name, attendance, comment):
    ts = datetime.now(YAKUTSK_TZ).strftime("%Y-%m-%d %H:%M:%S")
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO guests (name, attendance, comment, created_at) VALUES (?,?,?,?)",
            (name, attendance, comment, ts),
        )
        return cur.lastrowid, ts

def all_guests():
    with db() as conn:
        return conn.execute("SELECT * FROM guests ORDER BY id").fetchall()

# ─────────────────────────── Telegram helpers ───────────────────────────
def tg(method, **params):
    try:
        r = requests.post(f"{API}/{method}", json=params, timeout=20)
        return r.json()
    except Exception as e:
        print("Telegram error:", e)
        return {}

def tg_send(chat_id, text, reply_markup=None):
    params = {"chat_id": chat_id, "text": text, "parse_mode": "HTML"}
    if reply_markup:
        params["reply_markup"] = reply_markup
    return tg("sendMessage", **params)

def main_menu_markup():
    return {
        "inline_keyboard": [
            [{"text": "📥 Скачать список гостей (Excel)", "callback_data": "export"}],
            [{"text": "📊 Статистика", "callback_data": "stats"}],
        ]
    }

def notify_admins_new(guest_id, name, attendance, comment, ts):
    text = (
        "💌 <b>Новая заявка с сайта!</b>\n\n"
        f"<b>ФИО:</b> {name}\n"
        f"<b>Придёт:</b> {ATTENDANCE_RU.get(attendance, attendance)}\n"
        f"<b>Комментарий:</b> {comment or '—'}\n\n"
        f"<i>№{guest_id} · {ts}</i>"
    )
    for chat_id in ADMIN_CHAT_IDS:
        tg_send(chat_id, text, reply_markup=main_menu_markup())

# ─────────────────────────── Excel выгрузка ───────────────────────────
def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Гости"

    headers = ["№", "ФИО", "Сможете прийти", "Комментарий", "Дата отправки"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="7B1A3A")
    head_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for g in all_guests():
        ws.append([
            g["id"],
            g["name"],
            ATTENDANCE_RU.get(g["attendance"], g["attendance"]).replace("✅ ", "").replace("❌ ", "").replace("🤔 ", ""),
            g["comment"] or "",
            g["created_at"],
        ])

    widths = [6, 32, 22, 40, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def send_excel(chat_id):
    rows = all_guests()
    if not rows:
        tg_send(chat_id, "Пока нет ни одной заявки 🤷")
        return
    buf = build_excel()
    fname = f"guests_{datetime.now(YAKUTSK_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
    try:
        requests.post(
            f"{API}/sendDocument",
            data={"chat_id": chat_id, "caption": f"Список гостей · всего заявок: {len(rows)}"},
            files={"document": (fname, buf,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=60,
        )
    except Exception as e:
        print("sendDocument error:", e)
        tg_send(chat_id, "Не удалось сформировать файл 😔")

def send_stats(chat_id):
    rows = all_guests()
    total = len(rows)
    yes   = sum(1 for g in rows if g["attendance"] == "yes")
    maybe = sum(1 for g in rows if g["attendance"] == "maybe")
    no    = sum(1 for g in rows if g["attendance"] == "no")
    tg_send(chat_id,
        "📊 <b>Статистика заявок</b>\n\n"
        f"Всего: <b>{total}</b>\n"
        f"✅ Придут: <b>{yes}</b>\n"
        f"🤔 Не уверены: <b>{maybe}</b>\n"
        f"❌ Не смогут: <b>{no}</b>",
        reply_markup=main_menu_markup())

# ─────────────────────────── Бот: long polling ───────────────────────────
def polling_loop():
    print("Telegram polling запущен…")
    offset = None
    while True:
        try:
            r = requests.get(f"{API}/getUpdates",
                             params={"timeout": 30, "offset": offset}, timeout=40).json()
        except Exception as e:
            print("getUpdates error:", e)
            time.sleep(3)
            continue

        for upd in r.get("result", []):
            offset = upd["update_id"] + 1

            # обычные сообщения / команды
            msg = upd.get("message")
            if msg:
                chat_id = msg["chat"]["id"]
                text = (msg.get("text") or "").strip()
                if text in ("/start", "/menu"):
                    tg_send(chat_id,
                        "Привет! 👋\nЯ собираю заявки гостей со свадебного сайта "
                        "<b>vmwed.ru</b>.\n\nВыберите действие:",
                        reply_markup=main_menu_markup())
                else:
                    tg_send(chat_id, "Откройте меню командой /menu",
                            reply_markup=main_menu_markup())

            # нажатия инлайн-кнопок
            cq = upd.get("callback_query")
            if cq:
                chat_id = cq["message"]["chat"]["id"]
                data = cq.get("data")
                tg("answerCallbackQuery", callback_query_id=cq["id"])
                if data == "export":
                    send_excel(chat_id)
                elif data == "stats":
                    send_stats(chat_id)

# ─────────────────────────── Flask API ───────────────────────────
app = Flask(__name__)

@app.after_request
def cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = ALLOW_ORIGIN
    resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return resp

@app.route("/health")
def health():
    return jsonify(ok=True)

@app.route("/api/rsvp", methods=["POST", "OPTIONS"])
def rsvp():
    if request.method == "OPTIONS":
        return ("", 204)

    data = request.get_json(silent=True) or request.form
    name = (data.get("name") or "").strip()
    attendance = (data.get("attendance") or "").strip()
    comment = (data.get("comment") or "").strip()

    if not name or attendance not in ATTENDANCE_RU:
        return jsonify(ok=False, error="Заполните ФИО и выберите вариант ответа"), 400

    guest_id, ts = save_guest(name, attendance, comment)
    notify_admins_new(guest_id, name, attendance, comment, ts)
    return jsonify(ok=True, id=guest_id)

# ─────────────────────────── main ───────────────────────────
if __name__ == "__main__":
    if not BOT_TOKEN or not ADMIN_CHAT_IDS:
        print("⚠️  Задайте BOT_TOKEN и ADMIN_CHAT_IDS (см. .env.example)")
    init_db()
    # бот — в фоновом потоке, Flask — в основном
    threading.Thread(target=polling_loop, daemon=True).start()
    app.run(host="0.0.0.0", port=PORT)
